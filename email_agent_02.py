import smtplib
import imaplib
import csv
import os
import time
import random
import uuid
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.image import MIMEImage
from dotenv import load_dotenv
from openai import OpenAI
from docx import Document
from openpyxl import load_workbook

# Load environment variables
load_dotenv()

def get_ai_subject_line(client, job_title):
    """Generates professional subject line for Rec2Rec outreach."""
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "You are a professional recruiter. Output ONLY the subject line text."},
                {"role": "user", "content": f"Write a concise email subject line for Rec2Rec outreach about their '{job_title}' opening. Focus on partnership. No quotes."}
            ],
            max_tokens=30,
            temperature=0.7
        )
        return response.choices[0].message.content.strip().replace('"', '')
    except Exception as e:
        print(f"OpenAI API Error: {e}")
        return f"Partnership regarding {job_title}"

def read_template(file_path):
    """Reads text from .txt, .docx, or .html file."""
    clean_path = file_path.replace('"', '').replace("'", "").strip()
    if not os.path.exists(clean_path):
        raise FileNotFoundError(f"File not found: {clean_path}")
    
    if clean_path.lower().endswith('.docx'):
        doc = Document(clean_path)
        return '\n'.join([para.text for para in doc.paragraphs]), 'plain'
    elif clean_path.lower().endswith('.txt'):
        with open(clean_path, 'r', encoding='utf-8') as f:
            return f.read(), 'plain'
    elif clean_path.lower().endswith('.html'):
        with open(clean_path, 'r', encoding='utf-8') as f:
            return f.read(), 'html'
    else:
        raise ValueError(f"Unsupported format. Use .txt, .docx, or .html: {clean_path}")

def read_contacts_file(file_path):
    """Reads contacts from .csv or .xlsx file."""
    clean_path = file_path.replace('"', '').replace("'", "").strip()
    if not os.path.exists(clean_path):
        raise FileNotFoundError(f"File not found: {clean_path}")
    
    rows = []
    
    if clean_path.lower().endswith('.csv'):
        with open(clean_path, mode='r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            reader.fieldnames = [h.lower() for h in reader.fieldnames]
            if not any(x in reader.fieldnames for x in ['email', 'email id', 'email_id']):
                raise ValueError("CSV must contain an 'Email' column.")
            rows = [row for row in reader]
            
    elif clean_path.lower().endswith('.xlsx'):
        wb = load_workbook(clean_path, data_only=True)
        ws = wb.active
        
        headers = [cell.value.lower() if cell.value else '' for cell in ws[1]]
        
        if not any(x in headers for x in ['email', 'email id', 'email_id']):
            raise ValueError("Excel file must contain an 'Email' column.")
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):
                row_dict = dict(zip(headers, row))
                rows.append(row_dict)
    else:
        raise ValueError(f"Unsupported format. Use .csv or .xlsx: {clean_path}")
    
    return rows

def personalize_body(template, name, job_title, country):
    """Replaces placeholders in template - does NOT add signature."""
    body = template
    
    # Normalize NAME placeholders (case-insensitive)
    body = body.replace("[NAME]", name)
    body = body.replace("[Name]", name)
    body = body.replace("[name]", name)
    
    # Normalize JOB TITLE placeholders (case-insensitive)
    body = body.replace("[JOB TITLE]", job_title)
    body = body.replace("[Job Title]", job_title)
    body = body.replace("[job title]", job_title)
    body = body.replace("[JOB POSITION]", job_title)
    body = body.replace("[Job Position]", job_title)
    body = body.replace("[job position]", job_title)
    body = body.replace("[POSITION]", job_title)
    body = body.replace("[Position]", job_title)
    body = body.replace("[position]", job_title)
    
    # Normalize COUNTRY placeholders (case-insensitive)
    body = body.replace("[COUNTRY]", country)
    body = body.replace("[Country]", country)
    body = body.replace("[country]", country)
    
    # Return body without adding extra signature (it's already in the template)
    return body

def send_email_godaddy(server, sender, recipient, subject, body, content_type='plain', logo_path=None, sender_name="Zainab"):
    """Sends email with spam-prevention headers. Supports both plain text and HTML with inline logo."""
    msg = MIMEMultipart('related')
    msg['From'] = f"{sender_name} <{sender}>"
    msg['To'] = recipient
    msg['Subject'] = subject
    
    # CRITICAL: Spam prevention headers
    msg['Message-ID'] = f"<{uuid.uuid4()}@{sender.split('@')[1]}>"
    msg['Reply-To'] = sender
    
    # Add body (plain or HTML)
    msg.attach(MIMEText(body, content_type))
    
    # Attach logo inline if provided and content is HTML
    if content_type == 'html' and logo_path and os.path.exists(logo_path):
        try:
            with open(logo_path, 'rb') as f:
                img = MIMEImage(f.read())
                img.add_header('Content-ID', '<company_logo>')
                img.add_header('Content-Disposition', 'inline', filename=os.path.basename(logo_path))
                msg.attach(img)
        except Exception as e:
            print(f"      └─ ⚠️  Warning: Could not attach logo: {e}")
    
    server.send_message(msg)
    
    return msg  # Return message object for saving to Sent folder

def save_to_sent_folder(email_user, email_pass, msg, retries=3):
    """Saves a copy of the sent email to the Sent folder via IMAP with retry logic."""
    for attempt in range(retries):
        try:
            imap = imaplib.IMAP4_SSL('imap.secureserver.net', 993)
            imap.login(email_user, email_pass)
            
            # Try different common Sent folder names
            sent_folders = ['Sent', 'INBOX.Sent', '[Gmail]/Sent Mail', 'Sent Items']
            
            # List all folders to find the correct Sent folder
            status, folders = imap.list()
            
            # Try to append to Sent folder
            success = False
            for folder_name in sent_folders:
                try:
                    imap.append(folder_name, '\\Seen', None, msg.as_bytes())
                    success = True
                    break
                except:
                    continue
            
            if not success:
                # If standard names don't work, try 'Sent'
                imap.append('Sent', '\\Seen', None, msg.as_bytes())
            
            imap.logout()
            return True
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1)  # Wait 1 second before retry
                continue
            else:
                print(f"      └─ ⚠️  Could not save to Sent folder after {retries} attempts: {e}")
                return False
    
    return False

def main():
    print("=== AI Email Automation Agent (Rec2Rec) ===")
    print("NOTICE: Ensure your domain has SPF/DKIM configured to avoid spam!")

    # 1. Load Config
    api_key = os.getenv("OPENAI_API_KEY")
    email_user = os.getenv("EMAIL_USER")
    email_pass = os.getenv("EMAIL_PASSWORD")
    smtp_server = os.getenv("SMTP_SERVER", "smtpout.secureserver.net")
    smtp_port = int(os.getenv("SMTP_PORT", 465))

    if not all([api_key, email_user, email_pass]):
        print("❌ Error: Missing credentials in .env file.")
        return

    client = OpenAI(api_key=api_key)

    # 2. Get File Paths
    print("\n[Step 1] File Setup")
    print("Tip: If files are in this folder, just type the name (e.g., contacts.csv)")
    
    csv_input = input("Enter Contacts filename (.csv or .xlsx): ").strip().replace('"', '')
    txt_input = input("Enter Email Body filename (.docx, .txt, or .html): ").strip().replace('"', '')

    # 3. Read Template
    try:
        template_content, content_type = read_template(txt_input)
        print(f"✓ Template loaded ({len(template_content)} characters) - Type: {content_type.upper()}")
        # Show detected placeholders
        if "[Name]" in template_content or "[NAME]" in template_content:
            print("  └─ Detected [Name] placeholder")
        if "[JOB TITLE]" in template_content or "[Job Title]" in template_content or "[Position]" in template_content:
            print("  └─ Detected [Job Title] placeholder")
        if "[Country]" in template_content or "[COUNTRY]" in template_content:
            print("  └─ Detected [Country] placeholder")
    except Exception as e:
        print(f"❌ Error reading template: {e}")
        return

    # 4. Ask for logo if HTML template
    logo_path = None
    if content_type == 'html':
        logo_input = input("Enter Logo filename (e.g., logo.png) or press Enter to skip: ").strip().replace('"', '')
        if logo_input:
            logo_path = logo_input.replace('"', '').replace("'", "").strip()
            if os.path.exists(logo_path):
                print(f"✓ Logo found: {logo_path}")
            else:
                print(f"⚠️  Warning: Logo file not found. Emails will be sent without logo.")
                logo_path = None

    # 5. Read Contacts
    try:
        rows_to_process = read_contacts_file(csv_input)
        print(f"✓ Found {len(rows_to_process)} contacts.")
    except Exception as e:
        print(f"❌ Error reading contacts: {e}")
        return

    if input("\nStart sending? (yes/no): ").lower() != 'yes':
        print("Aborted by user.")
        return

    # 6. Connect to SMTP
    print(f"\n[Step 2] Connecting to {smtp_server}...")
    try:
        server = smtplib.SMTP_SSL(smtp_server, smtp_port)
        server.login(email_user, email_pass)
        print("✓ Login successful!")
    except Exception as e:
        print(f"❌ SMTP Connection Failed: {e}")
        return

    # 7. Process Loop
    print("\n[Step 3] Processing Emails...")
    print("⚠️  Sending with 3-7 second delays to protect sender reputation.")
    count = 0
    failed_count = 0
    
    for i, row in enumerate(rows_to_process, 1):
        name = str(row.get('name', 'Hiring Manager')).strip()
        email = str(row.get('email') or row.get('email id') or row.get('email_id', '')).strip()
        job_title = str(row.get('job title') or row.get('position', 'the role')).strip()
        country = str(row.get('country', 'US')).strip()

        if not email or '@' not in email:
            print(f"\n[{i}/{len(rows_to_process)}] ⚠️  Skipping: Invalid email")
            continue

        print(f"\n[{i}/{len(rows_to_process)}] Processing: {email}")
        print(f"    Name: {name} | Job Title: {job_title} | Country: {country}")
        
        # Generate AI subject
        print(f"    > Generating subject...")
        ai_subject = get_ai_subject_line(client, job_title)
        print(f"      └─ Subject: \"{ai_subject}\"")

        # Personalize Body
        personalized_body = personalize_body(template_content, name, job_title, country)
        
        # Preview first line (strip HTML tags for preview if HTML)
        if content_type == 'html':
            import re
            preview_text = re.sub('<[^<]+?>', '', personalized_body)[:100]
        else:
            preview_text = personalized_body.split('\n')[0]
        print(f"      └─ Body preview: {preview_text}...")

        try:
            msg_object = send_email_godaddy(server, email_user, email, ai_subject, personalized_body, content_type, logo_path)
            print(f"      └─ ✓ SENT")
            
            # Save to Sent folder
            if save_to_sent_folder(email_user, email_pass, msg_object):
                print(f"      └─ ✓ Saved to Sent folder")
            
            count += 1
        except smtplib.SMTPServerDisconnected:
            print(f"      └─ ⚠️  Connection lost. Reconnecting...")
            try:
                server.quit()
            except:
                pass
            
            # Reconnect
            try:
                server = smtplib.SMTP_SSL(smtp_server, smtp_port)
                server.login(email_user, email_pass)
                print(f"      └─ ✓ Reconnected successfully")
                
                # Retry sending this email
                msg_object = send_email_godaddy(server, email_user, email, ai_subject, personalized_body, content_type, logo_path)
                print(f"      └─ ✓ SENT (after reconnect)")
                
                if save_to_sent_folder(email_user, email_pass, msg_object):
                    print(f"      └─ ✓ Saved to Sent folder")
                
                count += 1
            except Exception as e:
                print(f"      └─ ❌ Reconnection FAILED: {e}")
                failed_count += 1
                continue
        except Exception as e:
            print(f"      └─ ❌ FAILED: {e}")
            failed_count += 1
            continue
        
        # Random delay (3-7 seconds)
        delay = random.uniform(3, 7)
        print(f"      └─ Waiting {delay:.1f}s...")
        time.sleep(delay)

    try:
        server.quit()
    except:
        pass
    
    print(f"\n✅ Done! Successfully sent {count}/{len(rows_to_process)} emails.")
    if failed_count > 0:
        print(f"⚠️  {failed_count} emails failed to send.")
    print("📧 Check your inbox for bounces and replies.")

if __name__ == "__main__":
    main()